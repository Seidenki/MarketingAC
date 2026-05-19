import json
import pandas as pd
from pathlib import Path

DATA_DIR = Path(__file__).parent.parent / "data"


def guardar_json(datos: list[dict], nombre_archivo: str) -> Path:
    DATA_DIR.mkdir(exist_ok=True)
    ruta = DATA_DIR / f"{nombre_archivo}.json"
    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(datos, f, ensure_ascii=False, indent=2)
    return ruta


def guardar_csv(datos: list[dict], nombre_archivo: str) -> Path:
    DATA_DIR.mkdir(exist_ok=True)
    ruta = DATA_DIR / f"{nombre_archivo}.csv"
    df = pd.DataFrame(datos)
    df["tipos"] = df["tipos"].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)
    df["horario"] = df["horario"].apply(lambda x: json.dumps(x, ensure_ascii=False) if isinstance(x, dict) else x)
    df.to_csv(ruta, index=False, encoding="utf-8-sig")
    return ruta


def guardar_excel(datos: list[dict], nombre_archivo: str) -> Path:
    DATA_DIR.mkdir(exist_ok=True)
    ruta = DATA_DIR / f"{nombre_archivo}.xlsx"
    
    df = pd.DataFrame(datos)
    df["tipos"] = df["tipos"].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)
    df["horario"] = df["horario"].apply(lambda x: json.dumps(x, ensure_ascii=False) if isinstance(x, dict) else x)
    
    df.columns = [c.upper() for c in df.columns]
    
    with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Resultados")
        
        ws = writer.sheets["Resultados"]
        
        # encabezados en negrita y con fondo
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # ajustar ancho de columnas automáticamente
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    
    return ruta

def exportar(datos: list[dict], nombre_archivo: str, formato: str = "json") -> Path:
    if formato == "json":
        return guardar_json(datos, nombre_archivo)
    elif formato == "csv":
        return guardar_csv(datos, nombre_archivo)
    elif formato == "excel":
        return guardar_excel(datos, nombre_archivo)
    else:
        raise ValueError(f"Formato no soportado: {formato}. Usa 'json', 'csv' o 'excel'.")